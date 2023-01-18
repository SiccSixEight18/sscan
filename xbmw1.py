import openpyxl

wb = openpyxl.load_workbook('xbmw.xlsx')

ws = wb['Sheet2']

gads1 = input("Ieraksti gadu:  ")
modelis1 = input("Ieraksti modeli:  ")

gads = ws.cell(row=3, column=3)
gads.value = gads1
gads.number_format = '0'

modelis = ws.cell(row=3, column=5)
modelis.value = modelis1
modelis.number_format = '0'

cena = ws.cell(3, 9)

cena1 = ws.cell(4, 9)

cena1.value = str(cena)

print(cena1)

wb.save('xbmw.xlsx')


#talak dabut third party softu lai var uztaisit interface no sita
#izdzest data_only vins dzes ara visas formulas wtflamo
